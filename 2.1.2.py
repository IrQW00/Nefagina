import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt


def do_exit(message: str):
    print(message)
    exit()


class UserInput:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")


class DataSet:
    def csv_reader(file_name: str):
        vacancies_array = []
        is_empty_file = True
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            if not reader.fieldnames:
                do_exit("Пустой файл")
            for row in reader:
                is_empty_file = False
                if all(row.values()):
                    vacancy = Vacancy(row)
                    vacancies_array.append(vacancy)
        if is_empty_file:
            do_exit("Нет данных")
        return vacancies_array


class Salary:
    salary_from: str
    salary_to: str
    salary_currency: str
    average_salary: int
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def set_attribute(self, key: str, value: str):
        if key == "salary_gross":
            return
        if key != 'salary_currency':
            value = float(value)
        self.__setattr__(key, value)

    def set_medianSalary(self):
        self.average_salary = int(self.currency_to_rub[self.salary_currency] *
                                  (float(self.salary_from) + float(self.salary_to)) // 2)


class Vacancy:
    name: str
    employer_name: str
    salary: Salary
    area_name: str
    published_at: int

    def __init__(self, fields: dict):
        self.date_time_publishing = None
        for key, value in fields.items():
            value = self.tag_refactor(value)
            if not self.check_salary(key, value):
                self.__setattr__(key, value)

        self.published_time_formatter()
        self.salary.set_medianSalary()

    def tag_refactor(self, value: str) -> str:
        value = re.sub('<.*?>', '', str(value)).replace("\r\n", "\n")
        value = ' '.join(value.split()).strip()
        return value

    def check_salary(self, key: str, value: str) -> bool:
        if not key.__contains__("salary"):
            return False
        if not hasattr(self, "salary"):
            self.salary = Salary()
        self.salary.set_attribute(key, value)
        return True

    def published_time_formatter(self):
        hour, minute, second = self.published_at.split('T')[1].split('+')[0].split(':')
        year, month, day = self.published_at.split('T')[0].split('-')
        self.date_time_publishing = datetime(int(year), int(month), int(day), int(hour), int(minute), int(second))
        self.published_at = int(year)

    def get_field(self, field: str):
        if field == 'salary':
            return self.salary.average_salary
        return self.__getattribute__(field)


class DataDictionary:
    def __init__(self):
        self.salary_years = {}
        self.vacancies_years = {}
        self.salary_years_by_profession = {}
        self.vacancies_years_by_profession = {}
        self.salaries_cities = {}
        self.vacancy_cities_ratio = {}
        self.city_vacancies_count = {}

    def update_data(self, vacancies: list, profession: str) -> None:
        self.profession = profession
        for vacancy in vacancies:
            self.update_data_by_vacancy(vacancy, profession)

        self.correct_data(vacancies)

    def update_data_by_vacancy(self, vacancy, profession: str):
        self.update_vacancy_dict('city_vacancies_count', 'area_name', vacancy)
        self.update_salary_dict('salary_years', 'published_at', vacancy)
        self.update_vacancy_dict('vacancies_years', 'published_at', vacancy)
        self.update_salary_dict('salaries_cities', 'area_name', vacancy)
        self.update_vacancy_dict('vacancy_cities_ratio', 'area_name', vacancy)
        if vacancy.name.__contains__(profession):
            self.update_salary_dict('salary_years_by_profession', 'published_at', vacancy)
            self.update_vacancy_dict('vacancies_years_by_profession', 'published_at', vacancy)

    def update_salary_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = [vac.salary.average_salary, 1]
        else:
            dictionary[key][0] += vac.salary.average_salary
            dictionary[key][1] += 1

    def update_vacancy_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = 1
        else:
            dictionary[key] += 1

    def correct_data(self, vacancies: list):
        for key, value in self.vacancy_cities_ratio.items():
            self.vacancy_cities_ratio[key] = round(value / len(vacancies), 4)

        buf = dict(sorted(self.salaries_cities.items(), key=lambda x: x[1][1] / x[1][0]))
        self.salaries_cities = self.get_first(buf, vacancies, 10)

        buf = dict(sorted(self.vacancy_cities_ratio.items(), key=lambda x: x[1], reverse=True))
        self.vacancy_cities_ratio = self.get_first(buf, vacancies, 10)

    def get_first(self, dictionary: dict, vacancies: list, amount: int) -> dict:
        count = 0
        res = {}
        for key, value in dictionary.items():
            if count == amount:
                break
            if self.city_vacancies_count[key] >= len(vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def print(self) -> None:
        print_dictionary: {str, dict} = {
            "Динамика уровня зарплат по годам: ": self.salary_years,
            "Динамика количества вакансий по годам: ": self.vacancies_years,
            "Динамика уровня зарплат по годам для выбранной профессии: ": self.salary_years_by_profession,
            "Динамика количества вакансий по годам для выбранной профессии: ": self.vacancies_years_by_profession,
            "Уровень зарплат по городам (в порядке убывания): ": self.salaries_cities,
            "Доля вакансий по городам (в порядке убывания): ": self.vacancy_cities_ratio
        }
        for key, value in print_dictionary.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            print(f"{key}{value}")


class Report:
    wb: Workbook

    def __init__(self, data: DataDictionary):
        self.work_book = Workbook()
        self.data = data

    def generate_excel(self):
        self.work_book.remove(self.work_book.active)
        self.generate_statistics_by_years()
        self.generate_statistics_by_cities()
        self.work_book.save("report.xlsx")

    def generate_statistics_by_years(self):
        ws = self.work_book.create_sheet("Статистика по годам")
        self.create_data_dict(ws, "A", "Год", {v: str(k) for k, v in data.salary_years.items()})
        self.create_data_dict(ws, "B", "Средняя зарплата", data.salary_years)
        self.create_data_dict(ws, "C", f"Средняя зарплата - {data.profession}", data.salary_years_by_profession)
        self.create_data_dict(ws, "D", "Количество вакансий", data.vacancies_years)
        self.create_data_dict(ws, "E", f"Количество вакансий - {data.profession}", data.vacancies_years_by_profession)
        self.update_cell_settings(ws)

    def create_data_dict(self, ws, column: str, name: str, dictionary: dict):
        ws[f"{column}1"] = name
        count = 2
        for year, value in dictionary.items():
            ws[f"{column}{count}"] = value
            count += 1

    def generate_statistics_by_cities(self):
        ws = self.work_book.create_sheet("Статистика по городам")
        self.create_data_dict(ws, "A",
                                      "Город", {v: k for k, v in data.salaries_cities.items()})
        self.create_data_dict(ws, "B", "Уровень зарплат", data.salaries_cities)
        self.create_data_dict(ws, "D",
                                      "Город", {v: k for k, v in data.vacancy_cities_ratio.items()})
        self.create_data_dict(ws, "E",
                                      "Доля вакансий", data.vacancy_cities_ratio)
        self.set_percent_style(ws)
        self.update_cell_settings(ws)

    def set_percent_style(self, ws):
        for i in range(2, 12):
            ws[f"E{i}"].number_format = FORMAT_PERCENTAGE_00

    def update_cell_settings(self, ws):
        self.set_cell(ws)
        self.set_correctly_column_width(ws)

    def set_cell(self, ws):
        isFirstCell = True
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=Side(border_style="thin", color="000000"),
                                         left=Side(border_style="thin", color="000000"),
                                         right=Side(border_style="thin", color="000000"),
                                         bottom=Side(border_style="thin", color="000000"))
                    if isFirstCell:
                        cell.font = Font(bold=True)
            isFirstCell = False

    def set_correctly_column_width(self, ws):
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value

    def generate_image(self):
        fig, ax = plt.subplots(2, 2)
        self.create_vertical_graph(ax, data.salary_years, data.salary_years_by_profession,
                                        "Уровень зарплат по годам")
        self.create_vertical_graph(ax, data.vacancies_years, data.vacancies_years_by_profession,
                                        "Количество вакансий по годам")
        self.create_horizont_graph(ax, "Уровень зарплат по городам")
        self.create_circle_graph(ax, "Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()

    def create_vertical_graph(self, ax, values, values_profession, title: str):
        my_label = ""
        coord = 0
        if title == "Уровень зарплат по годам":
            my_label = "Средняя зарплата"
        elif title == "Количество вакансий по годам":
            my_label = "Количество вакансий"
            coord = 1

        x_coord = [i for i in range(0, len(data.salary_years.keys()))]

        ax[0, coord].bar([x - 0.2 for x in x_coord], values.values(), width=0.5,
                         label=my_label)
        ax[0, coord].bar([x + 0.2 for x in x_coord], values_profession.values(), width=0.5,
                         label=f"{my_label} {data.profession}")
        ax[0, coord].set_xticks(x_coord, values.keys())
        ax[0, coord].set_xticklabels(values.keys(), rotation='vertical', va='top',
                                     ha='center')
        ax[0, coord].tick_params(axis='both', labelsize=8)
        ax[0, coord].legend(fontsize=8)
        ax[0, coord].grid(True, axis='y')
        ax[0, coord].set_title(title)

    def create_horizont_graph(self, ax, title: str):
        ax[1, 0].invert_yaxis()
        ax[1, 0].tick_params(axis='both', labelsize=8)
        ax[1, 0].set_yticklabels(list(data.salaries_cities.keys()), fontsize=6, va='center', ha='right')
        ax[1, 0].barh(list(data.salaries_cities.keys()), list(data.salaries_cities.values()))
        ax[1, 0].grid(True, axis='x')
        ax[1, 0].set_title(title)

    def create_circle_graph(self, ax, title: str):
        otherRatio = 1 - sum((list(data.vacancy_cities_ratio.values())))
        data.vacancy_cities_ratio.update({'Другие': otherRatio})
        ax[1, 1].pie(list(data.vacancy_cities_ratio.values()),
                     labels=list(data.vacancy_cities_ratio.keys()), textprops={'fontsize': 6})
        ax[1, 1].axis('scaled')
        ax[1, 1].set_title(title)


user_input = UserInput()
vacancies_array = DataSet.csv_reader(user_input.file_name)

if len(vacancies_array) == 0:
    do_exit("Ничего не найдено")

data = DataDictionary()
data.update_data(vacancies_array, user_input.profession)
data.print()
report = Report(data)
report.generate_image()